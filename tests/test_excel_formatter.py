"""Tests for generated Excel report formatting."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.reporting.excel_formatter import (
    format_control_report,
)


def test_format_control_report_preserves_values_and_adds_formatting(
        tmp_path: Path,
) -> None:
    """Formatting should improve presentation without changing control data."""
    output_path = tmp_path / "report.xlsx"

    summary = pd.DataFrame(
        {
            "Metric": [
                "Overall Status",
                "Payments Tested",
            ],
            "Value": [
                "REVIEW",
                3,
            ],
        }
    )

    exceptions = pd.DataFrame(
        {
            "Control Area": ["PAYMENT"],
            "Record ID": ["PAY-1001"],
            "Exception Type": ["UNALLOCATED PAYMENT"],
            "Exception Amount": [2_000.00],
            "Source Status": ["REVIEW"],
            "Severity": ["HIGH"],
            "Classification": ["ACTION REQUIRED"],
            "Settlement Case ID": ["CASE-001"],
            "Related Records": ["INV-1001"],
            "Link Basis": ["PAYMENT ALLOCATION RECORD"],
        }
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        summary.to_excel(
            writer,
            sheet_name="Run Summary",
            index=False,
        )

        exceptions.to_excel(
            writer,
            sheet_name="Settlement Exceptions",
            index=False,
        )

    format_control_report(output_path)

    workbook = load_workbook(output_path)

    try:
            summary_sheet = workbook["Run Summary"]
            exception_sheet = workbook["Settlement Exceptions"]

            assert summary_sheet["A2"].value == "Overall Status"
            assert summary_sheet["B2"].value == "REVIEW"
            assert summary_sheet["A3"].value == "Payments Tested"
            assert summary_sheet["B3"].value == 3

            assert summary_sheet.freeze_panes == "A2"
            assert summary_sheet["B2"].fill.fill_type == "solid"
            assert summary_sheet["B2"].font.bold is True

            assert exception_sheet["A2"].value == "PAYMENT"
            assert exception_sheet["B2"].value == "PAY-1001"
            assert exception_sheet["D2"].value == 2_000.00
            assert exception_sheet["D2"].number_format == "$#,##0.00"

            assert exception_sheet.freeze_panes == "A2"

            assert (
                "tblSettlementExceptions"
                in exception_sheet.tables
            )

            assert exception_sheet["F2"].fill.fill_type == "solid"
            assert exception_sheet["G2"].fill.fill_type == "solid"
            assert exception_sheet["G2"].font.bold is True

    finally:
            workbook.close()


def test_format_control_report_handles_empty_exception_sheet(
    tmp_path: Path,
) -> None:
    """An all-clear run should format correctly without creating an empty table."""
    output_path = tmp_path / "report.xlsx"

    summary = pd.DataFrame(
        {
            "Metric": ["Overall Status"],
            "Value": ["PASS"],
        }
    )

    exceptions = pd.DataFrame(
        columns=[
            "Control Area",
            "Record ID",
            "Exception Type",
            "Exception Amount",
            "Source Status",
            "Severity",
            "Classification",
            "Settlement Case ID",
            "Related Records",
            "Link Basis",
        ]
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        summary.to_excel(
            writer,
            sheet_name="Run Summary",
            index=False,
        )

        exceptions.to_excel(
            writer,
            sheet_name="Settlement Exceptions",
            index=False,
        )

    format_control_report(output_path)

    workbook = load_workbook(output_path)

    try:
        summary_sheet = workbook["Run Summary"]
        exception_sheet = workbook["Settlement Exceptions"]

        assert summary_sheet["B2"].value == "PASS"
        assert summary_sheet["B2"].fill.fill_type == "solid"

        assert exception_sheet.max_row == 1
        assert len(exception_sheet.tables) == 0
        assert exception_sheet.freeze_panes == "A2"

    finally:
        workbook.close()


def test_format_control_report_raises_when_summary_sheet_missing(
    tmp_path: Path,
) -> None:
    """A malformed generated workbook should fail clearly."""
    output_path = tmp_path / "report.xlsx"

    exceptions = pd.DataFrame(
        {
            "Record ID": ["PAY-1001"],
        }
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        exceptions.to_excel(
            writer,
            sheet_name="Settlement Exceptions",
            index=False,
        )

    with pytest.raises(
        KeyError,
        match="missing Run Summary",
    ):
        format_control_report(output_path)


def test_format_control_report_raises_when_exception_sheet_missing(
    tmp_path: Path,
) -> None:
    """A missing exception worksheet should fail clearly."""
    output_path = tmp_path / "report.xlsx"

    summary = pd.DataFrame(
        {
            "Metric": ["Overall Status"],
            "Value": ["PASS"],
        }
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        summary.to_excel(
            writer,
            sheet_name="Run Summary",
            index=False,
        )

    with pytest.raises(
        KeyError,
        match="missing Settlement Exceptions",
    ):
        format_control_report(output_path)