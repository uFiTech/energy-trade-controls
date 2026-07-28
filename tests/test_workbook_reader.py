"""Tests for workbook ingestion functions."""

from pathlib import Path 

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from src.ingestion.workbook_reader import (
    list_workbook_sheets,
    map_workbook_tables,
    read_excel_table,
    )


def test_list_workbook_sheets_returns_sheet_names(tmp_path: Path) -> None:
    """A valid workbook should return its worksheet names in order."""
    workbook_path = tmp_path / "test_portfolio.xlsx"

    workbook = Workbook()
    workbook.active.title = "Trades"
    workbook.create_sheet("Invoices")
    workbook.save(workbook_path)
    workbook.close()

    result = list_workbook_sheets(workbook_path)

    assert result == ["Trades", "Invoices"]


def test_list_workbook_sheets_raises_when_file_missing(
        tmp_path: Path,
) -> None:
    """A missing workbook should raise a clear file-not-found error."""
    missing_path = tmp_path / "missing_portfolio.xlsx"

    with pytest.raises(FileNotFoundError, match="Workbook not found"):
        list_workbook_sheets(missing_path)


def test_map_workbook_tables_returns_table_locations(
        tmp_path: Path,
) -> None:
    """Structured table names should map to their worksheet names."""
    workbook_path = tmp_path / "test_portfolio.xlsx"

    workbook = Workbook()

    trades_sheet = workbook.active
    trades_sheet.title = "Trades"
    trades_sheet.append(["Trade ID", "Quantity"])
    trades_sheet.append(["TRD-1001", 100_000])
    trades_sheet.add_table(
        Table(
            displayName="tblTrades",
            ref="A1:B2",
        )
    )

    invoices_sheet = workbook.create_sheet("Invoices")
    invoices_sheet.append(["Invoice ID", "Invoice Total"])
    invoices_sheet.append(["INV-2001", 1_500_000])
    invoices_sheet.add_table(
        Table(
            displayName="tblInvoices",
            ref="A1:B2",
        )
    )

    workbook.save(workbook_path)
    workbook.close()

    result = map_workbook_tables(workbook_path)

    assert result == {
        "tblTrades": "Trades",
        "tblInvoices": "Invoices",
    }


def test_map_workbook_tables_returns_empty_mapping_when_no_tables(
        tmp_path: Path,
) -> None:
    """A workbook without structured tables should return an empty mapping."""
    workbook_path = tmp_path / "empty_tables.xlsx"

    workbook = Workbook()
    workbook.save(workbook_path)
    workbook.close()

    result = map_workbook_tables(workbook_path)

    assert result == {}


def test_map_workbook_tables_raises_when_file_missing(
        tmp_path: Path,
) -> None:
    """A missing workbook should raise a clear file-not-found error."""
    missing_path = tmp_path / "missing_portfolio.xlsx"

    with pytest.raises(FileNotFoundError, match="Workbook not found"):
        map_workbook_tables(missing_path)


def test_read_excel_table_returns_dataframe_records(
        tmp_path: Path,
) -> None:
    """A populated Excel table should become a DataFrame."""
    workbook_path = tmp_path / "test_portfolio.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Trades"

    worksheet.append(["Trade ID", "Quantity"])
    worksheet.append(["TRD-1001", 100_000])
    worksheet.append(["TRD-1002", 75_000])

    worksheet.add_table(
        Table(
            displayName="tblTrades",
            ref="A1:B3",
        )
    )

    workbook.save(workbook_path)
    workbook.close()

    result = read_excel_table(workbook_path, "tblTrades")

    assert list(result.columns) == ["Trade ID", "Quantity"]
    assert result.to_dict(orient="records") == [
        {
             "Trade ID": "TRD-1001",
             "Quantity": 100_000,
        },
        {
            "Trade ID": "TRD-1002",
            "Quantity": 75_000,
        },
    ]


def test_read_excel_table_preserves_header_only_table(
        tmp_path: Path,
) -> None:
    """A header-only table should return an empty DataFrame with columns."""
    workbook_path = tmp_path / "header_only.xlsx"
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Trades"

    worksheet.append(["Trade ID", "Quantity"])

    worksheet.add_table(
        Table(
            displayName="tblTrades",
            ref="A1:B1",
        )
    )

    workbook.save(workbook_path)
    workbook.close()

    result = read_excel_table(workbook_path, "tblTrades")

    assert result.empty
    assert list(result.columns) == ["Trade ID", "Quantity"]
    assert len(result) == 0


def test_read_excel_table_raises_when_table_missing(
        tmp_path: Path,
) -> None:
    """A missing structured table should raise a clear key error."""
    workbook_path = tmp_path / "missing_portfolio.xlsx"

    workbook = Workbook()
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(KeyError, match="Excel table not found"):
        read_excel_table(workbook_path, "tblTrades")


def test_read_excel_table_raises_when_workbook_missing(
        tmp_path: Path,
) -> None:
    """A missing workbook should raise a clear file-not-found error."""
    missing_path = tmp_path / "missing_portfolio.xlsx"

    with pytest.raises(FileNotFoundError, match="Workbook not found"):
        read_excel_table(missing_path, "tblTrades")