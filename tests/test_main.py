"""Tests for the end-to-end control runner."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.main import run_controls


def test_run_controls_builds_review_summary_and_output(
        tmp_path: Path,
        monkeypatch,
) -> None:
    """An active settlement exception should produce a REVIEW run."""
    payments = pd.DataFrame(
        {
            "Payment ID": ["PAY-1001"],
            "Payment Amount": [100.00],
        }
    )

    allocations = pd.DataFrame(
        {
            "Payment ID": ["PAY-1001"],
            "Invoice ID": ["INV-1001"],
            "Applied Amount": [60.00],
        }
    )

    invoices = pd.DataFrame(
        {
            "Invoice ID": ["INV-1001"],
            "Invoice Total": [100.00],
            "Outstanding Amount": [0.00],
            "Invoice Status": ["Open"],
        }
    )

    tables = {
        "tblPayments": payments,
        "tblPaymentAllocations": allocations,
        "tblInvoices": invoices,
    }

    def fake_read_excel_table(
            workbook_path: str | Path,
            table_name: str,
    ) -> pd.DataFrame:
        return tables[table_name].copy()

    monkeypatch.setattr(
        "src.main.read_excel_table",
        fake_read_excel_table,
    )

    output_path = tmp_path / "settlement_exceptions.xlsx"

    summary, report = run_controls(
        workbook_path=tmp_path / "input.xlsx",
        output_path=output_path,

    )

    assert summary["Overall Status"] == "REVIEW"
    assert summary["Payments Tested"] == 1
    assert summary["Invoices Tested"] == 1
    assert summary["Control Observations"] == 2
    assert summary["Settlement Cases"] == 1
    assert summary["Action-Required Cases"] == 1
    assert summary["Controlled-Exclusion Cases"] == 0

    assert len(report) == 2
    assert report["Settlement Case ID"].nunique() == 1

    assert output_path.exists()

    workbook = load_workbook(
        output_path,
        read_only=True,
    )

    try:
        assert workbook.sheetnames == [
            "Run Summary",
            "Settlement Exceptions",
        ]
    finally:
        workbook.close()


def test_run_controls_returns_pass_when_no_exceptions(
        tmp_path,
        monkeypatch,
) -> None:
    """A fully reconciled sttlement population should produce PASS."""
    payments = pd.DataFrame(
        {
            "Payment ID": ["PAY-1002"],
            "Payment Amount": [100.00],
        }
    )

    allocations = pd.DataFrame(
        {
            "Payment ID": ["PAY-1002"],
            "Invoice ID": ["INV-1002"],
            "Applied Amount": [100.00],
        }
    )

    invoices = pd.DataFrame(
        {
            "Invoice ID": ["INV-1002"],
            "Invoice Total": [100.00],
            "Outstanding Amount": [0.00],
            "Invoice Status": ["Paid"],
        }
    )

    tables = {
        "tblPayments": payments,
        "tblPaymentAllocations": allocations,
        "tblInvoices": invoices,
    }

    def fake_read_excel_table(
        workbook_path: str | Path,
        table_name: str,
    ) -> pd.DataFrame:
        return tables[table_name].copy()

    monkeypatch.setattr(
        "src.main.read_excel_table",
        fake_read_excel_table,
    )

    output_path = tmp_path / "settlement_exceptions.xlsx"

    summary, report = run_controls(
        workbook_path=tmp_path / "input.xlsx",
        output_path=output_path,
    )

    assert summary["Overall Status"] == "PASS"
    assert summary["Control Observations"] == 0
    assert summary["Settlement Cases"] == 0
    assert summary["Action-Required Cases"] == 0
    assert summary["Controlled-Exclusion Cases"] == 0

    assert report.empty
    assert output_path.exists()

