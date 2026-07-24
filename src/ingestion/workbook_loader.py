from pathlib import Path

from openpyxl import load_workbook



PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = PROJECT_ROOT / "data" / "input" / "portfolio.xlsx"


REQUIRED_SHEETS = [
    "Trades",
    "Movements",
    "Invoices",
    "Payments",
    "Invoice Charges",
    "Service Commitments",
    "Claims",
    "Laytime & Demurrage",
    "Accruals",
    "Inventory",
    "Counterparties",
    "Products",
]


def validate_workbook_exists(workbook_path: Path) -> None:
    """"Confirm that the workbook exists before opening it."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected an .xlsx workbook, got: {workbook_path.suffix}")


def load_excel_workbook(workbook_path: Path):
    """Load the Excel workbook in read-only mode."""
    validate_workbook_exists(workbook_path)

    return load_workbook(
        filename=workbook_path,
        read_only=True,
        data_only=True,
    )


def validate_required_sheets(sheet_names: list[str]) -> list[str]:
    """Return a list of missing required workbook sheets."""
    missing_sheets = []

    for sheet in REQUIRED_SHEETS:
        if sheet not in sheet_names:
            missing_sheets.append(sheet)

    return missing_sheets


def run_workbook_validation(workbook_path: Path) -> None:
    """Load the workbook and validate required lifecycle sheets."""
    workbook = load_excel_workbook(workbook_path)
    sheet_names = workbook.sheetnames

    missing_sheets = validate_required_sheets(sheet_names)

    print("Workbook validation started")
    print(f"Project root: {PROJECT_ROOT}")
    print(f"Workbook path: {workbook_path}")
    print(f"Total sheets found: {len(sheet_names)}")

    if missing_sheets:
        print("Validation result: FAIL")
        print("Missing required sheets:")

        for sheet in missing_sheets:
            print(f"    -{sheet}")

        raise SystemExit(1)

    print("Validation result: PASS")
    print("All required lifecycle sheets were found.")


if __name__ == "__main__":
    run_workbook_validation(WORKBOOK_PATH)




    

