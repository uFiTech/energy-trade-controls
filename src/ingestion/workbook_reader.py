"""Functions for safely reading the source portfolio workbook."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def list_workbook_sheets(workbook_path: str | Path) -> list[str]:
    """Rturn the workbook names contained in an Excel workbook.

    Args:
        workbook_path: Location of the source Excel workbook.

    Returns:
        The workbook's worksheet names.

    Raises:
        FileNotFoundError: If the workbook does not exist.
    """
    path = Path(workbook_path)

    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def map_workbook_tables(workbook_path: str | Path) -> dict[str, str]:
    """Return structured Excel table names and their worksheets.

    Args:
        workbook_path: Location of the source Excel workbook.

    Returns:
        A mapping whose keys are Excel table names and whose values are
        worksheet names.

    Raises:
        FileNotFoundError: If the supplied workbook does not exist.
        ValueError: If the same table name is found more than once.
     """
    path = Path(workbook_path)

    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=False,
    )

    try:
        table_locations: dict[str, str] = {}

        for worksheet in workbook.worksheets:
            for table in worksheet.tables.values():
                if table.name in table_locations:
                    raise ValueError(
                        f"Duplicate Excel table names found: {table.name}"
                    )

                table_locations[table.name] = worksheet.title

        return table_locations

    finally:
        workbook.close()



def read_excel_table(
        workbook_path: str | Path,
        table_name: str,
) -> pd.DataFrame:
    """Read one structured Excel table into a pandas DataFrame.
    
    Args:
        workbook_path: Location of source Excel workbook.
        table_name: Structured Excel table name to load.
        
    Returns:
        A DataFrame containing the table's records and column names.
        A header-only table returns an empty DataFrame with columns.
        
    Raises:
        FileNotFoundError: If the workbook does not exist.
        KeyError: If the requested Excel table does not exist.
        ValueError: If the table range contains no header row.
    """
    path = Path(workbook_path)

    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=True,
    )

    try:
        for worksheet in workbook.worksheets:
            if table_name not in worksheet.tables:
                continue

            table = worksheet.tables[table_name]

            min_column, min_row, max_column, max_row = range_boundaries(
                table.ref
            )

            rows = list(
                worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_column,
                    max_col=max_column,
                    values_only=True,
                )
            )

            if not rows:
                raise ValueError(
                    f"Excel table has no header row: {table_name}"
                )

            headers = list(rows[0])
            data_rows = rows[1:]

            return pd.DataFrame(
                data_rows,
                columns=headers,
            )

        raise KeyError(f"Excel table not found: {table_name}")
    finally:
        workbook.close()