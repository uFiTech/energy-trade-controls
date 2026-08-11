"""Formatting for generated Excel settlement-control reports."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT =Font(
    color="FFFFFF",
    bold=True,
)

PASS_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6E0B4",
)

REVIEW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFE699",
)

HIGH_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

MEDIUM_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFE699",
)

INFO_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

ACTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

EXCLUSION_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7E6E6",
)


def _style_header(worksheet) -> None:
    """Apply a consistent style to the first worksheet row."""
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

def _find_column(
        worksheet,
        header_name: str,
) -> int | None:
    """Return the column number for a named header."""
    for cell in worksheet[1]:
        if cell.value == header_name:
            return cell.column

    return None


def _set_column_widths(
        worksheet,
        widths: dict[str, float],
) -> None:
    """Set worksheet widths using header names."""
    for header_name, width in widths.items():
        column_number = _find_column(
            worksheet,
            header_name,
        )

        if column_number is None:
            continue

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = width


def _format_run_summary(worksheet) -> None:
    """Format the run-summary worksheet."""
    _style_header(worksheet)

    worksheet.freeze_panes = "A2"

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 22

    metric_column = _find_column(
        worksheet,
        "Metric",
    )

    value_column = _find_column(
        worksheet,
        "Value",
    )

    if (
        metric_column is None
        or value_column is None
    ):
        return

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        metric = worksheet.cell(
            row=row_number,
            column=metric_column,
        ).value

        value_cell = worksheet.cell(
            row=row_number,
            column=value_column,
        )

        if metric != "Overall Status":
            continue

        status = str(
            value_cell.value or ""
        ).strip().upper()

        value_cell.font = Font(bold=True)

        if status == "PASS":
            value_cell.fill = PASS_FILL

        elif status == "REVIEW":
            value_cell.fill = REVIEW_FILL


def _format_settlement_exceptions(
        worksheet,
) -> None:
    """Format the settlement-exception worksheet."""
    _style_header(worksheet)

    worksheet.freeze_panes = "A2"

    _set_column_widths(
        worksheet,
        {
            "Control Area": 18,
            "Record ID": 17,
            "Exception Type": 27,
            "Exception Amount": 23,
            "Source Status": 18,
            "Severity": 15,
            "Classification": 27,
            "Settlement Case ID": 27,
            "Related Records": 27,
            "Link Basis": 32,
        },
    )

    amount_column = _find_column(
        worksheet,
        "Exception Amount",
    )

    severity_column = _find_column(
        worksheet,
        "Severity",
    )

    classification_column = _find_column(
        worksheet,
        "Classification",
    )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        if amount_column is not None:
            worksheet.cell(
                row=row_number,
                column=amount_column,
            ).number_format = "$#,##0.00"

        if severity_column is not None:
            severity_cell = worksheet.cell(
                row=row_number,
                column=severity_column,
            )

            severity = str(
                severity_cell.value or ""
            ).strip().upper()

            if severity == "HIGH":
                severity_cell.fill = HIGH_FILL

            elif severity == "MEDIUM":
                severity_cell.fill = MEDIUM_FILL

            elif severity == "INFO":
                severity_cell.fill = INFO_FILL

        if classification_column is not None:
            classification_cell = worksheet.cell(
                row=row_number,
                column=classification_column,
            )

            classification = str(
                classification_cell.value or ""
            ).strip().upper()

            if classification == "ACTION REQUIRED":
                classification_cell.fill = ACTION_FILL
                classification_cell.font = Font(
                    bold=True,
                )

            elif classification == "CONTROLLED EXCLUSION":
                classification_cell.fill = EXCLUSION_FILL

    if worksheet.max_row > 1:
        last_column = get_column_letter(
            worksheet.max_column
        )

        table_ref = (
            f"A1:{last_column}"
            f"{worksheet.max_row}"
        )

        table_name = "tblSettlementExceptions"

        if table_name not in worksheet.tables:
            table = Table(
                displayName=table_name,
                ref=table_ref,
            )

            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            worksheet.add_table(table)


def format_control_report(
        workbook_path: str | Path,
) -> None:
    """Apply presentation formatting to a generated control workbook."""
    workbook_path = Path(workbook_path)

    workbook = load_workbook(
        workbook_path,
    )

    try:
        if "Run Summary" not in workbook.sheetnames:
            raise KeyError(
                "Generated workbook is missing Run Summary sheet."
            )

        if "Settlement Exceptions" not in workbook.sheetnames:
            raise KeyError(
                "Generated workbook is missing "
                "Settlement Exceptions sheet."
            )

        _format_run_summary(
            workbook["Run Summary"],
        )

        _format_settlement_exceptions(
            workbook["Settlement Exceptions"],
        )

        workbook.save(
            workbook_path,
        )

    finally:
        workbook.close()


