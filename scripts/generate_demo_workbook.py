"""Generate a synthetic workbook for the public settlement-controls demo."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parent.parent

DEMO_PATH = (
    PROJECT_ROOT
    / "data"
    / "demo"
    / "demo_portfolio.xlsx"
)


def add_table_sheet(
        workbook: Workbook,
        sheet_name: str,
        table_name: str,
        headers: list[str],
        rows: list[list[object]],
) -> None:
    """Create a worksheet and add its records as a structured Excel table."""
    worksheet = workbook.create_sheet(sheet_name)

    worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    last_column = get_column_letter(len(headers))
    last_row = len(rows) + 1

    table = Table(
        displayName=table_name,
    ref=f"A1:{last_column}{last_row}",
    )

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"

    for column_number, header in enumerate(headers, start=1):
        values = [
            str(worksheet.cell(row=row_number, column=column_number).value or "")
            for row_number in range(1, last_row + 1)
        ]

        width = min(
            max(
                len(header),
                *(len(value) for value in values),
            )
            + 2,
            40,
        )

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = width


def build_demo_workbook() -> Workbook:
    """Build a workbook containing synthetic settlement scenarios."""
    workbook = Workbook()

    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(default_sheet)

    payments = [
        [
            "PAY-D001",
            100_000.00,
            "USD",
            "Clean settlement.",
        ],
        [
            "PAY-D002",
            80_000.00,
            "USD",
            "Partial allocation leaves 30,000 unapplied.",
        ],
        [
            "PAY-D003",
            50_000.00,
            "USD",
            "Deliberate overallocation control scenario.",
        ],
    ]

    add_table_sheet(
        workbook=workbook,
        sheet_name="Payments",
        table_name="tblPayments",
        headers=[
            "Payment ID",
            "Payment Amount",
            "Currency",
            "Notes",
        ],
        rows=payments
    )

    allocations = [
        [
            "PA-D001",
            "PAY-D001",
            "INV-D001",
            "TRD-D001",
            100_000.00,
        ],
        [
            "PA-D002",
            "PAY-D002",
            "INV-D002",
            "TRD-D002",
            50_000.00,
        ],
        [
            "PA-D003",
            "PAY-D003",
            "INV-D003",
            "TRD-D003",
            55_000.00,
        ],
    ]

    add_table_sheet(
        workbook=workbook,
        sheet_name="Payment Allocations",
        table_name="tblPaymentAllocations",
        headers=[
            "Allocation ID",
            "Payment ID",
            "Invoice ID",
            "Trade ID",
            "Applied Amount",
        ],
        rows=allocations,
    )

    invoices = [
        [
            "INV-D001",
            "TRD-D001",
            100_000.00,
            0.00,
            "Paid",
            "Clean settlement.",
        ],
        [
            "INV-D002",
            "TRD-D002",
            80_000.00,
            0.00,
            "Open",
            "Reported balance intentionally set to zero "
            "despite a 30,000 residual.",
        ],
        [
            "INV-D003",
            "TRD-D003",
            50_000.00,
            0.00,
            "Paid",
            "Allocation intentionally exceeds invoice by 5,000."
        ],
        [
            "INV-D004",
            "TRD-D004",
            15_000.00,
            0.00,
            "Cancelled",
            "Cancelled duplicate retained as a controlled exclusion.",
        ],  
        [
            "INV-D005",
            "TRD-D005",
            20_000.00,
            20_000.00,
            "Open",
            "Legitimate open invoice with correctly reported balance.",
        ],
    ]

    add_table_sheet(
        workbook=workbook,
        sheet_name="Invoices",
        table_name="tblInvoices",
        headers=[
             "Invoice ID",
            "Trade ID",
            "Invoice Total",
            "Outstanding Amount",
            "Invoice Status",
            "Notes",
        ],
        rows=invoices,
    )

    return workbook


def main() -> None:
    """Generate and save the public demo workbook."""
    workbook = build_demo_workbook()

    DEMO_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(DEMO_PATH)

    print(f"Demo workbook written to: {DEMO_PATH}")

if __name__ == "__main__":
    main()
       


